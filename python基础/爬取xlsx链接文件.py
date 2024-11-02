import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import json
import time
import random
from captcha import *
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import wait
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
from selenium import webdriver
from selenium.webdriver.support.select import Select

# 设置文件存放的地址
os.chdir(r'C:\\Users\\CYH10\\Desktop')

# 读取数据
link_date = load_workbook(r'C:\Users\CYH10\Desktop\数据链接.xlsx')
sheet = link_date.active
# 获取数据行数
rows = sheet.max_row
# 遍历数据
for i in range(rows):
    # 获取链接
    link = sheet.cell(row=i+1, column=1).value
    print(link)
    print('程序开始时间：', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    # 模拟使用edge浏览器登陆
    driver = webdriver.Edge()

    driver.get("https://www.dhsprogram.com/data/dataset_admin/login_main.cfm")
    time.sleep(5)  # 等待页面加载

    # 输入用户名和密码并点击登录
    username_input = driver.find_element(By.XPATH, "//input[@name='UserName']")
    username_input.send_keys("CYH107015@outlook.com")
    password_input = driver.find_element(By.XPATH, "//input[@name='UserPass']")
    password_input.send_keys("caoyihan7015")
    login_button = driver.find_element(By.XPATH, "//input[@type='submit']")
    login_button.click()

    ele = driver.find_element(By.XPATH, "//select[@name='proj_id']")
    select_ele = Select(ele)
    select_ele.select_by_value("206842")

    DM_button = driver.find_element(By.XPATH, "//input[@value='Download Manager']")
    DM_button.click()
    time.sleep(3)

    frame_1 = driver.find_element(By.XPATH, "//input[@name='ctrycode']")
    frame_1.click()
    frame_2 = driver.find_element(By.XPATH, "//input[@name='FileDataTypeCode']")
    frame_2.click()
    frame_3 = driver.find_element(By.XPATH, "//input[@value='he']")
    frame_3.click()
    frame_4 = driver.find_element(By.XPATH, "//input[@value='dt']")
    frame_4.click()
    # frame_5 = driver.find_element(By.XPATH, "//input[@name='suverymode']")
    # frame_5.click()
    build_list = driver.find_element(By.XPATH, "//input[@type='submit']")
    build_list.click()

    # 打开搜索页
    driver.get(link)
    time.sleep(10)  # 暂停20s
