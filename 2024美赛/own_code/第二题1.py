import numpy as np
import matplotlib.pyplot as plt
from scipy import stats
import random
import math
import seaborn as sns
from openpyxl import load_workbook

np.random.seed(43)
total_t = 100

i = stats.norm.rvs(5, 2, 1)  # 随机持续时间
print(i)

a = np.zeros(total_t)
a[0] = stats.norm.rvs(650, 116, 1)  # 随机生态位

m_1 = np.zeros(total_t)
f_1 = np.zeros(total_t)
m_2 = np.zeros(total_t)
f_2 = np.zeros(total_t)
m_1[0] = 10
f_1[0] = 10
m_2[0] = 10
f_2[0] = 10

for t in range(1, total_t):
    if t < i:
        a[t] = a[t - 1]
        m_1[t] = m_1[t - 1] + 0.6 * m_1[t - 1] * (1 - (m_1[t - 1] + 0.65 * f_1[t - 1]) / a[t])
        f_1[t] = f_1[t - 1] + 0.5 * f_1[t - 1] * (1 - (0.9 * m_1[t - 1] + f_1[t - 1]) / a[t])
        m_2[t] = m_2[t - 1] + 0.95 * m_2[t - 1] * (1 - (m_2[t - 1] + f_2[t - 1]) / a[t])
        f_2[t] = f_2[t - 1] + 0.85 * f_2[t - 1] * (1 - (m_2[t - 1] + f_2[t - 1]) / a[t])
        if m_1[t]< 0:
            m_1[t] = 0
        if f_1[t] < 0:
            f_1[t] = 0
        if m_2[t] < 0:
            m_2[t] = 0
        if f_2[t] < 0:
            f_2[t] = 0
    if t >= i:
        p = 1-math.exp(-0.05*i)
        b = random.uniform(0, 1)
        if p > b:
            a[t] = stats.norm.rvs(650, 116, 1)
            i = stats.norm.rvs(5, 2, 1) + t
            m_1[t] = m_1[t - 1] + 0.6 * m_1[t - 1] * (1 - (m_1[t - 1] + 0.65 * f_1[t - 1]) / a[t])
            f_1[t] = f_1[t - 1] + 0.5 * f_1[t - 1] * (1 - (0.9 * m_1[t - 1] + f_1[t - 1]) / a[t])
            m_2[t] = m_2[t - 1] + 0.95 * m_2[t - 1] * (1 - (m_2[t - 1] + f_2[t - 1]) / a[t])
            f_2[t] = f_2[t - 1] + 0.85 * f_2[t - 1] * (1 - (m_2[t - 1] + f_2[t - 1]) / a[t])
            if m_1[t] < 0:
                m_1[t] = 0
            if f_1[t] < 0:
                f_1[t] = 0
            if m_2[t] < 0:
                m_2[t] = 0
            if f_2[t] < 0:
                f_2[t] = 0
        else:
            a[t] = a[t - 1]
            i = stats.norm.rvs(5, 2, 1) + t
            m_1[t] = m_1[t - 1] + 0.6 * m_1[t - 1] * (1 - (m_1[t - 1] + 0.65 * f_1[t - 1]) / a[t])
            f_1[t] = f_1[t - 1] + 0.5 * f_1[t - 1] * (1 - (0.9 * m_1[t - 1] + f_1[t - 1]) / a[t])
            m_2[t] = m_2[t - 1] + 0.95 * m_2[t - 1] * (1 - (m_2[t - 1] + f_2[t - 1]) / a[t])
            f_2[t] = f_2[t - 1] + 0.85 * f_2[t - 1] * (1 - (m_2[t - 1] + f_2[t - 1]) / a[t])
            if m_1[t] < 0:
                m_1[t] = 0
            if f_1[t] < 0:
                f_1[t] = 0
            if m_2[t] < 0:
                m_2[t] = 0
            if f_2[t] < 0:
                f_2[t] = 0
b = []
b.append(a.T)
sns.heatmap(b, cmap='viridis', annot=True, cbar=False)
fig, ax = plt.subplots()

ax.plot(m_1, label='lampreys_m')
ax.plot(f_1, label='lampreys_f')
ax.imshow(b, extent=ax.get_xlim() + ax.get_ylim(), aspect='auto', origin='lower', alpha=0.5, cmap='viridis')
ax.legend()

fig2, ax = plt.subplots()
ax.plot(m_2, label='other population_m')
ax.plot(f_2, label='other population_f')
ax.imshow(b, extent=ax.get_xlim() + ax.get_ylim(), aspect='auto', origin='lower', alpha=0.5, cmap='viridis')
ax.legend()

plt.show()

print(m_1)
print(f_1)
print(m_2)
print(f_2)

xlsx = load_workbook("C:\\Users\\CYH10\Desktop\\不同环境.xlsx")
sheet = xlsx.active
for c in range(len(a)):
    sheet.cell(row=c + 1, column=1).value = a[c]

for c in range(len(m_1)):
    sheet.cell(row=c + 1, column=2).value = m_1[c]

for d in range(len(f_1)):
    sheet.cell(row=d + 1, column=3).value = f_1[d]

for e in range(len(m_2)):
    sheet.cell(row=e + 1, column=4).value = m_2[e]

for f in range(len(f_2)):
    sheet.cell(row=f + 1, column=5).value = f_2[f]

xlsx.save("C:\\Users\\CYH10\Desktop\\不同环境.xlsx")