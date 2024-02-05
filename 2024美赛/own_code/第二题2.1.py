import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import expon, norm
from scipy import stats
import random
import math
from openpyxl import load_workbook

total_t = 100

m1_total = []
m2_total = []
f1_total = []
f2_total = []

m_1 = np.zeros(total_t)
f_1 = np.zeros(total_t)
m_2 = np.zeros(total_t)
f_2 = np.zeros(total_t)
m_1[0] = 10
f_1[0] = 10
m_2[0] = 10
f_2[0] = 10

for t in range(1, total_t):
    m_1[t] = m_1[t - 1] + 0.6 * m_1[t - 1] * (1 - (m_1[t - 1] + 0.65 * f_1[t - 1]) / 300)
    f_1[t] = f_1[t - 1] + 0.5 * f_1[t - 1] * (1 - (0.9 * m_1[t - 1] + f_1[t - 1]) / 300)
    m_2[t] = m_2[t - 1] + 0.95 * m_2[t - 1] * (1 - (m_2[t - 1] + f_2[t - 1]) / 300)
    f_2[t] = f_2[t - 1] + 0.85 * f_2[t - 1] * (1 - (m_2[t - 1] + f_2[t - 1]) / 300)
    if m_1[t]< 0:
        m_1[t] = 0
    if f_1[t] < 0:
        f_1[t] = 0
    if m_2[t] < 0:
        m_2[t] = 0
    if f_2[t] < 0:
        f_2[t] = 0

population1_result = np.zeros(total_t)
population2_result = np.zeros(total_t)

for i in range(total_t):
    population1_result[i] = m_1[i] + f_1[i]
    population2_result[i] = m_2[i] + f_2[i]

plt.figure()
plt.plot(population1_result, label='population lampreys')
plt.plot(population2_result, label='population other')
plt.legend(loc='upper right')
plt.show()
print(population1_result)
print(population2_result)

xlsx = load_workbook("C:\\Users\\CYH10\Desktop\\不同环境.xlsx")
sheet = xlsx.active

for c in range(len(population1_result)):
    sheet.cell(row=c + 1, column=8).value = population1_result[c]

for d in range(len(population2_result)):
    sheet.cell(row=d + 1, column=9).value = population2_result[d]


xlsx.save("C:\\Users\\CYH10\Desktop\\不同环境.xlsx")