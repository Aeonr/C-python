import numpy as np
import matplotlib.pyplot as plt
from scipy import stats
import random
import math
import seaborn as sns
import pandas as pd
from openpyxl import load_workbook
import xlrd

np.random.seed(43)
total_t = 500

v = stats.norm.rvs(30, 4, 1)  # 随机持续时间

a = np.zeros(total_t)
a[0] = stats.norm.rvs(0.5, 0.1, 1)  # 随机

i = np.zeros(total_t)
j = np.zeros(total_t)
k = np.zeros(total_t)

i[0] = 10
j[0] = 10
k[0] = 10

for t in range(1, total_t):
    if t < v:
        a[t] = a[t - 1]
        i[t] = i[t-1] + 0.6 * i[t - 1] * (1 - (i[t - 1] - 0.5 * j[t - 1] + 0.3 * k[t - 1]) / 1000) - a[t] * i[t - 1] + math.cos(i[t - 1])
        j[t] = j[t-1] + 0.6 * j[t - 1] * (1 - (j[t - 1] - 0.5 * i[t - 1] + 0.4 * k[t - 1]) / 1000) - 0.5 * j[t - 1] + math.cos(j[t - 1])
        k[t] = j[t-1] + 0.6 * k[t - 1] * (1 - (k[t - 1] + 0.3 * i[t - 1] + 0.4 * j[t - 1]) / 1000) - 0.5 * k[t - 1] + math.cos(k[t - 1])
        if i[t] < 0:
            i[t] = 0
        if j[t] < 0:
            j[t] = 0
        if k[t] < 0:
            k[t] = 0
    if t >= v:
        p = 1-math.exp(-0.05*v)
        b = random.uniform(0, 1)
        if p > b:
            a[t] = stats.norm.rvs(0.5, 0.1, 1)
            v = stats.norm.rvs(30, 4, 1) + t
            i[t] = i[t - 1] + 0.6 * i[t - 1] * (1 - (i[t - 1] - 0.5 * j[t - 1] + 0.3 * k[t - 1]) / 1000) - a[t] * i[t - 1] + math.cos(i[t - 1])
            j[t] = j[t - 1] + 0.6 * j[t - 1] * (1 - (j[t - 1] - 0.5 * i[t - 1] + 0.4 * k[t - 1]) / 1000) - 0.5 * j[t - 1] + math.cos(j[t - 1])
            k[t] = j[t - 1] + 0.6 * k[t - 1] * (1 - (k[t - 1] + 0.3 * i[t - 1] + 0.4 * j[t - 1]) / 1000) - 0.5 * k[t - 1] + math.cos(k[t - 1])
            if i[t] < 0:
                i[t] = 0
            if j[t] < 0:
                j[t] = 0
            if k[t] < 0:
                k[t] = 0
        else:
            a[t] = a[t - 1]
            v = stats.norm.rvs(30, 4, 1) + t
            i[t] = i[t - 1] + 0.6 * i[t - 1] * (1 - (i[t - 1] - 0.5 * j[t - 1] + 0.3 * k[t - 1]) / 1000) - a[t] * i[t - 1] + math.cos(i[t - 1])
            j[t] = j[t - 1] + 0.6 * j[t - 1] * (1 - (j[t - 1] - 0.5 * i[t - 1] + 0.4 * k[t - 1]) / 1000) - 0.5 * j[t - 1] + math.cos(j[t - 1])
            k[t] = j[t - 1] + 0.6 * k[t - 1] * (1 - (k[t - 1] + 0.3 * i[t - 1] + 0.4 * j[t - 1]) / 1000) - 0.5 * k[t - 1] + math.cos(k[t - 1])
            if i[t] < 0:
                i[t] = 0
            if j[t] < 0:
                j[t] = 0
            if k[t] < 0:
                k[t] = 0
b = []
b.append(a.T)
sns.heatmap(b, cmap='viridis', annot=True, cbar=False)
fig, ax = plt.subplots()

ax.plot(i, label='i')
ax.plot(j, label='j')
ax.plot(k, label='k')
ax.imshow(b, extent=ax.get_xlim() + ax.get_ylim(), aspect='auto', origin='lower', alpha=0.5, cmap='viridis')
ax.legend()
print(a)
xlsx = load_workbook("C:\\Users\\CYH10\Desktop\\热力图.xlsx")
sheet = xlsx.active
for c in range(len(a)):
    sheet.cell(row=c + 1, column=1).value = a[c]

for c in range(len(i)):
    sheet.cell(row=c + 1, column=2).value = i[c]

for d in range(len(j)):
    sheet.cell(row=d + 1, column=3).value = j[d]

for e in range(len(k)):
    sheet.cell(row=e + 1, column=4).value = k[e]

xlsx.save("C:\\Users\\CYH10\Desktop\\热力图.xlsx")

print(i)
print(j)
print(k)
