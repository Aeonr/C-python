import numpy as np
import matplotlib.pyplot as plt
from scipy import stats
from openpyxl import load_workbook

x = np.arange(0.0, 1.1, 0.1)
d = np.arange(0, 2.1, 0.1)

result = np.zeros(4)
k = 0.5
result_rm = []
result_rf = []
result_a = []
result_b = []

m_best = []
f_best = []

for n in range(100, 501):
    for rm in x:
        for rf in x:
            for a in d:
                for b in d:
                    m = np.zeros(n)
                    f = np.zeros(n)
                    m[0] = 10
                    f[0] = 10
                    for t in range(1, n):
                        m[t] = m[t - 1] + rm * m[t - 1] * (1 - (m[t - 1] + a * f[t - 1])/1000)
                        f[t] = f[t - 1] + rf * f[t - 1] * (1 - (b * m[t - 1] + f[t - 1])/1000)
                        if t == n - 1:
                            e = m[n-1] / (m[n-1] + f[n-1])
                            if abs(e - 0.56) < k:
                                result[0] = rm
                                result[1] = rf
                                result[2] = a
                                result[3] = b
                                m_best = m
                                f_best = f
                                k = abs(e - 0.56)
                            else:
                                continue
    result_rm.append(result[0])
    result_rf.append(result[1])
    result_a.append(result[2])
    result_b.append(result[3])
    print(n)

xlsx = load_workbook("C:\\Users\\CYH10\Desktop\\灵敏度分析1.xlsx")
sheet = xlsx.active
for c in range(401):
    sheet.cell(row=c + 1, column=1).value = 100 + c

for c in range(len(result_rm)):
    sheet.cell(row=c + 1, column=2).value = result_rm[c]

for d in range(len(result_rf)):
    sheet.cell(row=d + 1, column=3).value = result_rf[d]

for e in range(len(result_a)):
    sheet.cell(row=e + 1, column=4).value = result_a[e]

for f in range(len(result_b)):
    sheet.cell(row=f + 1, column=5).value = result_b[f]

xlsx.save("C:\\Users\\CYH10\Desktop\\灵敏度分析1.xlsx")

plt.figure()
plt.plot([100, 501], result_rm, label='result_rm')
plt.plot([100, 501], result_rf, label='result_rf')
plt.plot([100, 501], result_a, label='result_a')
plt.plot([100, 501], result_b, label='result_b')

plt.legend()
plt.show()
