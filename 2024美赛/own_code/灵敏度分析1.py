import numpy as np
import matplotlib.pyplot as plt
from scipy import stats
from openpyxl import load_workbook

x = np.arange(0.0, 1.1, 0.1)
d = np.arange(0, 2.1, 0.1)
total_t = 100
result = np.zeros(4)
k = 0.5
result_rm = []
result_rf = []
result_a = []
result_b = []

m_best = []
f_best = []
env = stats.norm.rvs(650, 116, 100)

for i in range(100):
    for rm in x:
        for rf in x:
            for a in d:
                for b in d:
                    m = np.zeros(total_t)
                    f = np.zeros(total_t)
                    m[0] = 10
                    f[0] = 10
                    for t in range(1, total_t):
                        m[t] = m[t - 1] + rm * m[t - 1] * (1 - (m[t - 1] + a * f[t - 1])/env[i])
                        f[t] = f[t - 1] + rf * f[t - 1] * (1 - (b * m[t - 1] + f[t - 1])/env[i])
                        if t == total_t - 1:
                            e = m[total_t-1] / (m[total_t-1] + f[total_t-1])
                            if abs(e + (22/70000) * env[i] - 122/700) < k:
                                result[0] = rm
                                result[1] = rf
                                result[2] = a
                                result[3] = b
                                m_best = m
                                f_best = f
                                k = abs(e - 0.56)
                            else:
                                continue
    result_rm.append(result[0])
    result_rf.append(result[1])
    result_a.append(result[2])
    result_b.append(result[3])
    print(i)

xlsx = load_workbook("C:\\Users\\CYH10\Desktop\\灵敏度分析.xlsx")
sheet = xlsx.active
for c in range(len(env)):
    sheet.cell(row=c + 1, column=1).value = env[c]

for c in range(len(result_rm)):
    sheet.cell(row=c + 1, column=2).value = result_rm[c]

for d in range(len(result_rf)):
    sheet.cell(row=d + 1, column=3).value = result_rf[d]

for e in range(len(result_a)):
    sheet.cell(row=e + 1, column=4).value = result_a[e]

for f in range(len(result_b)):
    sheet.cell(row=f + 1, column=5).value = result_b[f]

xlsx.save("C:\\Users\\CYH10\Desktop\\灵敏度分析.xlsx")

plt.figure()
plt.plot(env, result_rm, label='result_rm')
plt.plot(env, result_rf, label='result_rf')
plt.plot(env, result_a, label='result_a')
plt.plot(env, result_b, label='result_b')

plt.legend()
plt.show()

# import pandas as pd
# import matplotlib.pyplot as plt
# import numpy as np
#
# data = pd.read_excel("C:\\Users\\CYH10\\Desktop\\灵敏度分析.xlsx")
#
# env = data['env']
# rm = data['rm']
# rf = data['rf']
# a = data['a']
# b = data['b']
#
# p1 = np.polyfit(env, rm, deg=1)
# plt.plot(env, np.polyval(p1, env), label="rm")
# p2 = np.polyfit(env, rf, deg=1)
# plt.plot(env, np.polyval(p2, env), label="rf")
# p3 = np.polyfit(env, a, deg=1)
# plt.plot(env, np.polyval(p3, env), label="a")
# p4 = np.polyfit(env, b, deg=1)
# plt.plot(env, np.polyval(p4, env), label="b")
# plt.legend()
# plt.show()
